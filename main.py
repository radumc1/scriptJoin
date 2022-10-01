from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
import auxiliar
# workbook = load_workbook(filename="2_bun.xlsx")
# sheet = workbook.active
#
# workbook1 = load_workbook(filename="1_bunx3.xlsx")
# sheet1 = workbook1.active
# # print(sheet[f'AT2'].value)
# # for row in sheet[1:100]:
# #     for item in row[1:10]:
# #         print(item.value)
#    # for value in row:
#
#      # print(value.value)
#
# # for row in sheet.iter_rows(min_row=1, max_col=3, max_row=2626):
# #      for cell in row:
# #          print(cell.value)
#
# # for x in range(1,2625):
# #        for y in range(1,101):
# #            sheet.cell(row=x, column=y)
#
# for i in range(1, sheet.max_row + 1):
#     # print("\n")
#     # print("Row ", i, " data :")
#     for j in range(1, sheet.max_column + 1):
#         if(get_column_letter(j)=="A"):
#             cell_obj = sheet.cell(row=i, column=j)
#             print(cell_obj.value)
#
#         # print("COLOANA ESTE",get_column_letter(j),cell_obj.value, end=" ")

import pandas as pd
customers = pd.read_excel("1_bunx3.xlsx")
calls = pd.read_excel("2_bun.xlsx")
outer_join_df = customers.merge(calls, how="outer", on="Parcela")
outer_join_df.to_excel("OuterJoin.xlsx")